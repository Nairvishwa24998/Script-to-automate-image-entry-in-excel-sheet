# Import relevant directories
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


# function to obtain image from file and add it at the appropriate position in the sheet
def image_loader(file, counter):
    img = Image(file)
    img.anchor = "B" + str(counter)
    # Dimensions of the image needed
    img.height = 95.039
    img.width = 101.7599
    ws.add_image(img)

wb = load_workbook(filename)
ws = wb.active


# Go through each cell of the appropriate row
for entry in ws['D']:
    for file in os.listdir(filelocation):
        # Only consider the file if its of type jpg
        if file.endswith('.jpg'):
            if (len(str(entry.value).lower().strip()) <= len(file[:-4].lower().strip()) and str(entry.value).lower().strip() in file[:-4].lower().strip()) or (len(str(entry.value).lower().strip()) > len(file[:-4].lower().strip()) and file[:-4].lower().strip() in str(entry.value).lower().strip()):
                counter = entry.row
                # Use the function to add the image
                image_loader(file,counter)


wb.save(filename)
wb.close()
