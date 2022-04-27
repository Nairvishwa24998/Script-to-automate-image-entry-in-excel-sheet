import os
from openpyxl import load_workbook

# Obtain the name of the excel file
filename = input()
filelocation = input()

# Loading workbook
wb = load_workbook(filename)
ws = wb.active
counter = 2

# iterate through the files in the location and add the filenames of files with jpg. extension as entries in the sheet
for files in os.listdir(filelocation):
    if files.endswith('jpg'):
        for entry in ws['D']:
                ws.cell(row = counter, column= 4).value = files[:-4]
        counter += 1



wb.save(filename)

